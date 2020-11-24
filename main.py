import openpyxl
import datetime
from datetime import date
import random
first = (
    "Super", "Retarded", "Great", "Sexy", "Vegan", "Brave", "Shy", "Cool", "Poor", "Rich", "Fast", "Gummy", "Yummy",
    "Masked", "Unusual", "American", "Bisexual", "MLG", "Mlg", "lil", "Lil")
second = (
    "Coder", "Vegan", "Man", "Hacker", "Horse", "Bear", "Goat", "Goblin", "Learner", "Killer", "Woman", "Programmer",
    "Spy",
    "Stalker", "Spooderman", "Carrot", "Goat", "Quickscoper", "Quickscoper")


def initialize_scheduler(sheet):
    """Initializing the sheet"""

    sheet['B1'] = 'Track1'
    sheet['C1'] = 'Track2'
    sheet['D1'] = 'Track3'
    sheet['E1'] = 'Track4'

    sheet['B2'] = 'Room P1 & P2'
    sheet['C2'] = 'Room P3'
    sheet['D2'] = 'Room P4'
    sheet['E2'] = 'Room P5'

    return sheet


def merge_cells(sheet):
    """Merging the cells"""

    sheet.merge_cells('B3:E3')
    sheet['A3'] = '8:00-'
    sheet['B3'] = '                                 Registration                              '
    sheet['A4'] = '9:00 - 10:15'
    sheet.merge_cells('B4:E4')
    sheet['B4'] = '                 Welcome & Keynote (Sydney Resnick) (Room P1 and P2)             '
    sheet['A5'] = '10:15 - 10:45'
    sheet.merge_cells('B5:E5')
    sheet['B5'] = '                                     Morning Tea                                          '
    sheet['A6'] = 'Session 1 \n' \
                  '10:45 - 12:15'


def chair_demand(events):
    print("do not chair the lecture which the chair wants to be a part of")
    for key in events:
        chairs = events[key]
        print(key, "lecture\n")
        for index in range(0, len(chairs)):
            print(index, chairs[index], '\n')  # print for user and select the chair

        while True:
            ind = input("Kindly select the chair that wants to attend the lecture\n"
                        "if no chair is selected press 0 ")

            if ind == '0':
                break
            else:
                chairs.pop(ind)


def assign_chair(lectures):
    events = {}
    lecturers = {}
    chairs = []
    lec_updated = []
    for lec in lectures:
        lec_updated.append(' '.join(lec))

    for lec in lec_updated:
        index = 0
        if lec == '':
            continue
        while index < 4:
            # chairs.append(input("kindly enter the chair\n"))
            firrst = random.choice(first)
            seccond = random.choice(second)
            chairs.append((firrst + " " + seccond))
            if index < 1:
                firrst = random.choice(first)
                seccond = random.choice(second)
                lecturers[str(lec)] = str(firrst) + " " + str(seccond)
                print(lecturers[str(lec)])
                is_absent = input("--------If the lecturer is absent press 0 otherwise press any other KEY------------")
                if is_absent == '0':
                    break
                is_first = input(
                    "--------------If the lecturer is a first time speaker kindly"
                    " press 0 otherwise press anyother key---------------\n")
                if is_first == '0':
                    mentor = input("kindly enter the name of the mentor that will be assigned to the lecturer\n")
                    chairs.append(mentor)
                    index += 1
            index += 1
        if str(lec) not in events:
            print("Chair entered to lectures :  " + str(lec) + "\n")
            events[str(lec)] = chairs[:]
            chairs.clear()

    return lec_updated, events, lecturers


def read_event_file():
    """This function reads file and assign chairs and lecturers to it"""
    print("-----------------------------Reading Event File-------------------------------")
    file1 = open('events.txt', 'r')
    lines = file1.readlines()
    lectures = []
    for line in lines:
        line = line.strip('\n')
        lecture = line.split(' ')
        lectures.append(lecture[:-1])

    lec_updated, events, lecturers = assign_chair(lectures)

    return lec_updated, events, lecturers


def get_priority_events(str_list):
    priority_events = []
    while True:
        print("\nKindly select which event would you like to priorotize\n")
        for index in range(0, len(str_list)):
            print('\n')
            print(index, " - " + str_list[index])
        print("\n kindly select the index given  to the lecturer to insert it to the priority queue\n")
        event_index = int(input("Kindly enter the index\n"))
        priority_val = str_list.pop(event_index)
        priority_events.append(priority_val)
        try:
            decision = int(input("\nDo you still want any other event to be given priority? \n"
                                 "\n If no then then press 0 else press any other Key \n"))
        except:
            while True:
                try:
                    print("\n------------------You did not enter any (index)"
                          "number try again please----------------------\n    ")
                    decision = int(input("\n Do you still want any other event in priority? \n"
                                         "\n If no then then press 0 else "
                                         "press any other Key \n"))
                    break
                except:
                    continue

        if decision == 0:
            break
        else:
            continue
    return priority_events


def input_from_user(events):
    str_list = list(filter(None, events))

    priority_events = get_priority_events(str_list)

    return priority_events


def same_audience(events):
    str_list = list(filter(None, events))
    # list of events of that will attract the same audience are organized in the same room
    same_aud = []
    while True:
        index = 0
        for index in range(0, len(str_list)):
            print('\n')
            print(index, str_list[index])
        print("conference organizer says that the audience for this talk is the same \n")
        print("kindly specify the events that you think will"
              " have the same audience so that they"
              " do not have to change the room ")
        event_index = int(input("please enter the value"))
        try:
            priority_val = str_list.pop(event_index)
            _ = events.pop(event_index)
        except ValueError:
            print("Kindly input the index in range")
            continue
        same_aud.append(priority_val)
        if len(same_aud) == 2:
            index = int(input("If all the similar audiences are identified press 0 "
                          " to terminate otherwise press any other Key"))

        if int(index) == 0:
            break

    return same_aud, events


def same_lecture(events):
    print("conference organizer says talk x y and z are on the same topic try not to schedule them on the same time")

    str_list = list(filter(None, events))
    # list of events of the same topic that should not be placed in the same session
    for index in range(0, len(events)):
        print(index, events[index])
    same_events = []

    while True:
        for index in range(0, len(str_list)):
            print('\n')
            print(index, str_list[index])
        event_index = int(input("please enter the value\n"))

        try:
            priority_val = str_list.pop(event_index)
            _ = events.pop(event_index)
        except:
            print("The index is out of range, kindly input the index inside the range ")
            priority_val = str_list.pop(event_index)
            _ = events.pop(event_index)

        same_events.append(priority_val)
        index = input("If all the similar events are identified press 0 to terminate the program")
        index = int(index)
        if int(index) == 0:
            break

    return same_events, events


def remove_from_list(events, first_row, second_row):
    print("remove an event from the list")
    for event in events:
        if event in first_row:
            events.remove(event
                          )
        elif event in second_row:
            events.remove(event)


def fill_scheduler(priority_events, events, same_events, same_aud):
    first_row = []
    second_row = []
    event_matrix = []
    episode = 0
    for event in priority_events:
        first_row.append(event)
        second_row.append(event)
        episode += 2

    while episode < 8:
        while True:
            if not same_aud:
                break
            event = same_aud.pop()
            first_row.append(event)

            episode += 1
            if episode < 8:
                event = same_aud.pop()
                second_row.append(event)
                episode += 1
            else:
                break
        while True:
            if not same_events and len(first_row) < 5:
                break
            first_row.append(same_events.pop())
            episode += 1
        while len(first_row) < 4:
            episode += 1
            first_row.append(events.pop())
        while len(second_row) < 4:
            episode += 1
            second_row.append(events.pop())

    event_matrix.append(first_row)
    event_matrix.append(second_row)

    return event_matrix


def input_days():

    while True:
        try:
            days = int(input("How many days are required by the user? \n"))
            break
        except ValueError:
            print("Kindly input days as an integer\n")
    print("---------------No of days input by user:\n-------------------" + str(days))
    while True:
        try:
            time_slots = int(input("How many time slots are required by the user? \n"))
            break
        except ValueError:
            print("Kindly input time_slots as an integer\n")
    if time_slots > 4:
        print("\n Maximum three time_slots can be placed in a day\n")
        while True:
            input(time_slots)
            if time_slots < 5:
                break

    return days, time_slots


def get_events_list(events, days):
    event_list = []
    days = int(days)
    ev_index = 0
    while days > 0:
        index = 0
        row = []
        while index < 8:
            row.append(events[ev_index])
            index += 1
            ev_index += 1
        event_list.append(row[:])
        row.clear()
        days += -1
    return event_list


def run():
    events, chairs, lecturers = read_event_file()
    days, time_slots = input_days()
    event_list = get_events_list(events, days)
    for events in event_list:
        priority_events = input_from_user(events)
        same_events, events = same_lecture(events)
        same_aud, events = same_audience(events)
        event_list = fill_scheduler(priority_events, events, same_events, same_aud)
        today = date.today()
        d1 = today.strftime("%d-%m-%Y")
        print("Today's date:", today)
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet = initialize_scheduler(sheet)
        merge_cells(sheet)
        file_name = str(d1) + ".xlsx"
        today += datetime.timedelta(days=1)

        wb.save(file_name)


if __name__ == '__main__':
    run()
